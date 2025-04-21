import os
from dotenv import load_dotenv
from googleapiclient.discovery import build
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

load_dotenv()

# Load the API key from the .env file
api_key = os.getenv("YOUTUBE_API_KEY")
if api_key is None:
    raise ValueError("API key not found. Please set the YOUTUBE_API_KEY environment variable.")

youtube = build("youtube",version="v3",developerKey=api_key)

#get the channel upload playlist id
channel_id = os.getenv("CHANNEL_ID")
if channel_id is None:
    raise ValueError("Channel ID not found. Please set the CHANNEL_ID environment variable.")
channel_response = youtube.channels().list(
    part="contentDetails",
    id = channel_id
).execute()

uploads_playlist_id = channel_response["items"][0]["contentDetails"]["relatedPlaylists"]["uploads"]

#get the videos
response = youtube.playlistItems().list(
    part="snippet",
    playlistId=uploads_playlist_id,
    maxResults=50
).execute()

#save the details to an excel sheet
video_data = []
for item in response["items"]:
    video_data.append({
        "Title": item["snippet"]["title"],
        "Video ID": item["snippet"]["resourceId"]["videoId"],
        "Thumbnail URL": item["snippet"]["thumbnails"]["default"]["url"],
        "Link": f"https://www.youtube.com/watch?v={item['snippet']['resourceId']['videoId']}"
    })
df = pd.DataFrame(video_data)
excel_file = "youtube_videos.xlsx"
sheet_name = "Videos"


# Check if the Excel file exists
if os.path.exists(excel_file):
    # string_to_use = string.ascii_letters + string.digits
    # sheet_name += ''.join(random.choice(string_to_use) for _ in range(8))
    book = load_workbook(excel_file)

    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
    
    #replace this sheet to be removed
    new_sheet = book.create_sheet("_temp_sheet_")
    book.remove(sheet)
    new_sheet.title = sheet_name
    sheet = book[sheet_name] # Get the new sheet object

    # Now write the DataFrame to the new sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, header=True, index=False)):
        sheet.append(list(row))

    # Save the changes to the workbook
    book.save(excel_file)
else:
    # New file
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

# Auto-adjust column widths
book = load_workbook(excel_file)
sheet = book[sheet_name]

for column_cells in sheet.columns:
    max_length = 0
    col = column_cells[0].column_letter
    for cell in column_cells:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = (max_length + 2)
    sheet.column_dimensions[col].width = adjusted_width

book.save(excel_file)
print(f"✅ Video details saved to {excel_file} with auto-adjusted columns.")
